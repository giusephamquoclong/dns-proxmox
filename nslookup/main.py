import requests
import re
import unidecode
import urllib3
import json
import os
import sys
from office365.runtime.auth.authentication_context import AuthenticationContext
from office365.sharepoint.client_context import ClientContext
from office365.sharepoint.files.file import File
from office365.runtime.auth.user_credential import UserCredential
from office365.sharepoint.files.file_system_object_type import FileSystemObjectType
import io
import pandas as pd
import openpyxl
import json
from colorama import Fore 
from termcolor import colored 
from pprint import PrettyPrinter, pprint
from operator import add, itemgetter, attrgetter
from pygments import highlight, lexers, formatters
import pymsteams

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


ip_proxmox='10.200.104.3'

def get_CSRF(ip):
    url = f"https://{ip}:8006/api2/json/access/ticket"
    headers = {'Content-Type': 'application/x-www-form-urlencoded',
               'Content-Type': 'application/json',
               'Connection': 'keep-alive'}
    login = {'username': 'root@pam',
             'password': 'Abc@1234'}
    resp = requests.post(url, headers=headers,
                         data=json.dumps(login), verify=False)
    # data=cookie.json()
    # print(data['response'])
    cookie = json.loads(resp.text)
    # print(cookie["data"]["ticket"])
    return cookie["data"]["CSRFPreventionToken"]

def get_cookie(ip):
    url=f"https://{ip}:8006/api2/json/access/ticket"
    headers = {'Content-Type': 'application/x-www-form-urlencoded',
               'Content-Type': 'application/json',
               'Connection':'keep-alive'}
    login = {'username':'root@pam',
             'password':'Abc@1234'}
    resp = requests.post(url,headers=headers,data=json.dumps(login),verify=False)
    # data=cookie.json()
    # print(data['response'])
    cookie=json.loads(resp.text)
    # print(cookie["data"]["ticket"])
    return cookie["data"]["ticket"]

def get_value_excel(url,filename,username,password):
# target url taken from sharepoint and credentials
    relative_url = filename
    ctx_auth = AuthenticationContext(url)
    if ctx_auth.acquire_token_for_user(username, password):
        ctx = ClientContext(url, ctx_auth)
        web = ctx.web
        ctx.load(web)
        ctx.execute_query()
        print("Authentication successful")
   
    response = File.open_binary(ctx, relative_url)

    # save data to BytesIO stream
    bytes_file_obj = io.BytesIO()
    bytes_file_obj.write(response.content)
    bytes_file_obj.seek(0) 
     
    # set file object to start
    excel = openpyxl.load_workbook(bytes_file_obj)
    sheet = excel.active
    m_row = sheet.max_row
    m_col = sheet.max_column
    dns=[]
    for i in range(1, m_col + 1):
        # cell = sheet.cell(row=1, column=i)
        for j in range(1,m_row):
            cell= sheet.cell(row=j,column=j)
            if j!=m_row and sheet.cell(row=j+1,column=i).value !=None:
                dns.append({"dns": sheet.cell(row=1,column=i).value,
                        "value": sheet.cell(row=j+1,column=i).value})
    return dns

def add_ip_ipset_cluster(ip,ip_ipset,domain,name_ipset):
    cookie = 'PVEAuthCookie='+str(get_cookie(ip))
    CSRF = str(get_CSRF(ip))
    url = "https://"+str(ip)+":8006/api2/json/cluster/firewall/ipset/"+name_ipset
    # print(url)
    headers = {'Content-Type': 'application/x-www-form-urlencoded',
               'Content-Type': 'application/json',
               'Connection': 'keep-alive',
               'Cookie': cookie,
               'CSRFPreventionToken': CSRF}
    notify=[]
    
    for i in ip_ipset:
        required = {'cidr': i, 'name': name_ipset, 'comment': domain}
        notify.append(required)

        resp = requests.post(url, headers=headers,
                             data=json.dumps(required), verify=False)
    return notify

def readfile(url):
    ip_file=[]
    with open(url) as fp:
        Lines = fp.readlines()
        for line in Lines:
            ip_file.append(line.strip())
        return(ip_file)
    
    
    
def check_ip_exist_in_ipset(ip,name_ipset):
    cookie='PVEAuthCookie='+str(get_cookie(ip))
    # print(cookie)
    url = f"https://{ip}:8006/api2/json/cluster/firewall/ipset/{name_ipset}"
    headers = {'Content-Type': 'application/x-www-form-urlencoded',
               'Content-Type': 'application/json',
               'Connection': 'keep-alive',
               'Cookie': cookie}
    # parameter = {'type': 'vm'}
    resp = requests.get(url, headers=headers,verify=False)
    datas = json.loads(resp.text)
    # vm = []
    # pprint(datas)
    ip_whitelistss=readfile("ipwhitelist.txt")
    ip_whitelists=[]
    for ip_wl in ip_whitelistss:
        ip_whitelists.append(ip_wl)
    
    # pprint(ip_whitelists)
    for data in datas['data']:
        for ip_whitelist in ip_whitelists:
            if ip_whitelist == data['cidr']:
                ip_whitelists.remove(ip_whitelist)
                # print(ip_whitelist)
                
    return ip_whitelists

def nslookup(ip,name_ipset):
    cookie = 'PVEAuthCookie='+str(get_cookie(ip))
    CSRF = str(get_CSRF(ip))
    name_ipset = 'linux-repo-crm'
    url = "https://"+str(ip)+":8006/api2/json/cluster/firewall/ipset/"+name_ipset
    headers = {'Content-Type': 'application/x-www-form-urlencoded',
               'Content-Type': 'application/json',
               'Connection': 'keep-alive',
               'Cookie': cookie,
               'CSRFPreventionToken': CSRF}
    notifys=[]
    domains=get_value_excel("https://3cxcloudonline.sharepoint.com/sites/IDBCorporation","/sites/IDBCorporation/Shared Documents/TeamX/url-dns.xlsx","bruce@idb.com.vn","Minhtam123")
    for domain in domains:
        # pprint(domain)
        cmd = os.system("nslookup "+ domain['value'] +" | grep Address | grep -v \"#53\" | awk '{print $2}' > ipwhitelist.txt")
        if check_ip_exist_in_ipset(ip,name_ipset) == []:
            # print("da co")
            continue
        else:
            ipset=check_ip_exist_in_ipset(ip,name_ipset)
            notify=add_ip_ipset_cluster(ip,ipset,domain['dns'],name_ipset)
            notifys.append(notify)
           
           
    pprint(notifys)
    myTeamsMessage = pymsteams.connectorcard("https://3cxcloudonline.webhook.office.com/webhookb2/7f7e9c6b-8e0f-4e40-9b3a-140874f72aaf@2c5396f8-3de5-4223-9e8d-0d244bedbeb5/IncomingWebhook/40c5e39ae3a4467a866f299c9acb39c2/25b4588d-f2c8-472a-99c7-1881aaec4ec5")
    myTeamsMessage.text(str(notifys))
    myTeamsMessage.send()
    
if __name__ == '__main__':
    nslookup(ip_proxmox,"linux-repo-crm")